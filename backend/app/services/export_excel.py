"""
Export Excel des données du Module Statistiques — écran "Statistiques",
bouton "Exporter en Excel". Un onglet par catégorie demandée
(commandes/paiements/passeports émis/contrôles), chacun filtrable par pays
et par année — jamais une seule feuille fourre-tout, pour rester exploitable
telle quelle (tri, filtre automatique Excel) sans retraitement.

Volontairement des LISTES DE LIGNES DÉTAILLÉES (une ligne par commande, par
paiement, ...) — pas les agrégats déjà disponibles à l'écran (voir
app.services.statistiques) : l'export sert à un usage comptable/d'audit qui
a besoin du détail, l'écran sert au pilotage qui a besoin de la synthèse.
"""
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import extract, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.commande import Commande
from app.models.controle import Controle
from app.models.paiement import Paiement
from app.models.passeport import Passeport, StatutPasseport
from app.models.pays import Pays
from app.models.utilisateur import Utilisateur

CATEGORIES_VALIDES = {"commandes", "paiements", "passeports_emis", "controles"}

_POLICE_ENTETE = Font(name="Arial", bold=True, color="FFFFFF")
_FOND_ENTETE = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid")
_POLICE_NORMALE = Font(name="Arial", size=10)


def _ecrire_feuille(classeur: Workbook, nom_feuille: str, entetes: list[str], lignes: list[list]) -> None:
    feuille = classeur.create_sheet(nom_feuille)
    feuille.append(entetes)
    for colonne in range(1, len(entetes) + 1):
        cellule = feuille.cell(row=1, column=colonne)
        cellule.font = _POLICE_ENTETE
        cellule.fill = _FOND_ENTETE
    for ligne in lignes:
        feuille.append(ligne)
        for colonne in range(1, len(ligne) + 1):
            feuille.cell(row=feuille.max_row, column=colonne).font = _POLICE_NORMALE
    feuille.freeze_panes = "A2"
    feuille.auto_filter.ref = f"A1:{get_column_letter(len(entetes))}{max(feuille.max_row, 1)}"
    for colonne in range(1, len(entetes) + 1):
        feuille.column_dimensions[get_column_letter(colonne)].width = 20


async def _noms_pays(db: AsyncSession) -> dict[int, str]:
    result = await db.execute(select(Pays))
    return {p.id: p.nom for p in result.scalars().all()}


async def generer_export_excel(
    db: AsyncSession, categories: set[str], pays_id: int | None = None, annee: int | None = None
) -> bytes:
    noms_pays = await _noms_pays(db)
    classeur = Workbook()
    classeur.remove(classeur.active)  # la feuille vierge par défaut — remplacée par nos propres onglets

    if "commandes" in categories:
        query = select(Commande)
        if pays_id is not None:
            query = query.where(Commande.pays_id == pays_id)
        if annee is not None:
            query = query.where(extract("year", Commande.cree_le) == annee)
        result = await db.execute(query)
        lignes = [
            [
                c.id[:8],
                noms_pays.get(c.pays_id, f"Pays #{c.pays_id}"),
                c.quantite,
                float(c.montant_total),
                c.statut.value,
                c.mode_impression.value,
                c.responsable_nom,
                c.cree_le.strftime("%Y-%m-%d") if c.cree_le else "",
            ]
            for c in result.scalars().all()
        ]
        _ecrire_feuille(
            classeur, "Commandes",
            ["ID", "Pays", "Quantité", "Montant (XAF)", "Statut", "Mode impression", "Responsable", "Date"],
            lignes,
        )

    if "paiements" in categories:
        query = select(Paiement, Commande.pays_id).select_from(Paiement).join(Commande, Paiement.commande_id == Commande.id)
        if pays_id is not None:
            query = query.where(Commande.pays_id == pays_id)
        if annee is not None:
            query = query.where(extract("year", Paiement.cree_le) == annee)
        result = await db.execute(query)
        lignes = [
            [
                p.id[:8],
                noms_pays.get(pid, f"Pays #{pid}"),
                float(p.montant),
                p.devise,
                p.moyen.value,
                p.statut.value,
                p.cree_le.strftime("%Y-%m-%d") if p.cree_le else "",
            ]
            for p, pid in result.all()
        ]
        _ecrire_feuille(classeur, "Paiements", ["ID", "Pays", "Montant", "Devise", "Moyen", "Statut", "Date"], lignes)

    if "passeports_emis" in categories:
        query = select(Passeport).where(Passeport.statut != StatutPasseport.PRECHARGE)
        if pays_id is not None:
            query = query.where(Passeport.pays_id == pays_id)
        if annee is not None:
            query = query.where(Passeport.numero_annee == str(annee))
        result = await db.execute(query)
        lignes = [
            [
                f"{p.numero_pays}-{p.numero_annee}-{p.numero_lot}",
                noms_pays.get(p.pays_id, f"Pays #{p.pays_id}"),
                p.statut.value,
            ]
            for p in result.scalars().all()
        ]
        _ecrire_feuille(classeur, "Passeports émis", ["Numéro PPB", "Pays", "Statut"], lignes)

    if "controles" in categories:
        query = (
            select(Controle, Passeport, Utilisateur)
            .join(Passeport, Controle.passeport_id == Passeport.id)
            .join(Utilisateur, Controle.agent_id == Utilisateur.id)
        )
        if pays_id is not None:
            query = query.where(Passeport.pays_id == pays_id)
        if annee is not None:
            query = query.where(Passeport.numero_annee == str(annee))
        result = await db.execute(query)
        lignes = [
            [
                f"{p.numero_pays}-{p.numero_annee}-{p.numero_lot}",
                noms_pays.get(p.pays_id, f"Pays #{p.pays_id}"),
                c.poste_id,
                c.resultat.value,
                agent.nom_complet,
                c.cree_le.strftime("%Y-%m-%d %H:%M") if c.cree_le else "",
            ]
            for c, p, agent in result.all()
        ]
        _ecrire_feuille(classeur, "Contrôles (passeports vérifiés)", ["Numéro PPB", "Pays", "Poste", "Résultat", "Agent", "Date"], lignes)

    if len(classeur.sheetnames) == 0:
        _ecrire_feuille(classeur, "Export", ["Aucune catégorie sélectionnée"], [])

    feuille_infos = classeur.create_sheet("Filtres appliqués", 0)
    feuille_infos.append(["Généré le", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    feuille_infos.append(["Pays", noms_pays.get(pays_id, "Tous les pays") if pays_id is not None else "Tous les pays"])
    feuille_infos.append(["Année", str(annee) if annee is not None else "Toutes les années"])
    feuille_infos.append(["Catégories", ", ".join(sorted(categories)) if categories else "Aucune"])
    for ligne_num in range(1, 5):
        feuille_infos.cell(row=ligne_num, column=1).font = Font(name="Arial", bold=True)
        feuille_infos.cell(row=ligne_num, column=2).font = _POLICE_NORMALE
    feuille_infos.column_dimensions["A"].width = 18
    feuille_infos.column_dimensions["B"].width = 30

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
